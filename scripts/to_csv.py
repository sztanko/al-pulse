import typer
from pathlib import Path
from openpyxl import load_workbook
import csv

app = typer.Typer()

@app.command()
def convert(
    input_dir: Path = typer.Argument(..., exists=True, file_okay=False),
    output_csv: Path = typer.Argument(...)
):
    all_rows = []
    header_written = False

    for xlsx_file in sorted(input_dir.glob("*.xlsx")):
        wb = load_workbook(xlsx_file, read_only=True)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if not header_written:
                    all_rows.append(row)
                    header_written = True
            else:
                all_rows.append(row)

    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(all_rows)

    typer.echo(f"Combined CSV written to: {output_csv}")

if __name__ == "__main__":
    app()
