from pathlib import Path
from datetime import datetime, timedelta
from calendar import monthrange
import time
import csv
import logging
import typer
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from tenacity import (
    retry,
    stop_after_attempt,
    retry_if_exception_type,
    wait_random_exponential,
)

# ---- Constants ----
URL = "https://rnt.turismodeportugal.pt/RNT/Pesquisa_AL.aspx"
DATE_INPUT_SELECTOR = "input[placeholder='AAAA-MM-DD']"
SEARCH_BUTTON_SELECTOR = "input[value='Pesquisar']"
EXPORT_LINK_TEXT = "Exportar detalhe registos"
DOWNLOAD_DIR = Path.cwd() / "downloads"
TEMP_FILENAME = "al_data.xlsx"
FIRST_DATE = "2007-01-01"
TIMEOUT_SECONDS = 7
OUTPUT_CSV = f"{DOWNLOAD_DIR}/al/al_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

# ---- Logging ----
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger(__name__)


def end_of_month(dt: datetime) -> datetime:
    last_day = monthrange(dt.year, dt.month)[1]
    return dt.replace(day=last_day)


END_OF_CURRENT_MONTH = end_of_month(datetime.now()).strftime("%Y-%m-%d")

app = typer.Typer(pretty_exceptions_enable=False)


def append_excel_to_csv(
    excel_path: Path, csv_path: Path, write_header: bool, ts: datetime
):
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    # Make sure the CSV directory exists
    # csb_path is the full path to the CSV file
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    count_writes = 0
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Write the timestamp as the first column
            if i == 0:
                if write_header:
                    row = ("etl_timestamp",) + row
                else:
                    continue
            else:
                row = (ts.strftime("%Y-%m-%d %H:%M:%S"),) + row
            writer.writerow(row)
            count_writes += 1
    return count_writes


@retry(
    stop=stop_after_attempt(5),
    wait=wait_random_exponential(multiplier=1.2, min=2, max=10),
    retry=retry_if_exception_type(Exception),
    reraise=True,
)
def fetch_and_export(
    page, from_date: str, to_date: str, csv_path: Path, write_header: bool, ts: datetime
):
    t1 = time.time()
    log.info(f"Fetching from {from_date} to {to_date}")

    page.goto(URL)

    date_inputs = page.locator(DATE_INPUT_SELECTOR)
    if date_inputs.count() < 2:
        raise RuntimeError("Expected 2 date inputs, found less.")
    date_inputs.nth(0).fill(from_date)
    date_inputs.nth(1).fill(to_date)

    page.locator(SEARCH_BUTTON_SELECTOR).click()
    page.wait_for_timeout(TIMEOUT_SECONDS * 1000)

    if not page.locator(f"a:has-text('{EXPORT_LINK_TEXT}')").is_visible(
        timeout=TIMEOUT_SECONDS * 1000
    ):
        log.warning(f"No data to export")
        return

    with page.expect_download() as download_info:
        page.locator(f"a:has-text('{EXPORT_LINK_TEXT}')").click()
    download = download_info.value

    DOWNLOAD_DIR.mkdir(exist_ok=True)
    save_path = DOWNLOAD_DIR / TEMP_FILENAME
    download.save_as(save_path)

    num_rows = append_excel_to_csv(
        save_path, csv_path, write_header=write_header, ts=ts
    )
    save_path.unlink()
    t2 = time.time()
    log.info(
        f"Fetched and saved {num_rows} rows for month {from_date[0:7]} in {(t2 - t1):.2f} seconds."
    )


@app.command()
def run_per_month(
    from_dt: str = typer.Argument(FIRST_DATE, help="Start date in YYYY-MM-DD"),
    to_dt: str = typer.Argument(END_OF_CURRENT_MONTH, help="End date in YYYY-MM-DD"),
    output: Path = typer.Option(OUTPUT_CSV, help="Output CSV path"),
):
    start_date = datetime.strptime(from_dt, "%Y-%m-%d")
    end_date = end_of_month(datetime.strptime(to_dt, "%Y-%m-%d"))
    if start_date > end_date:
        raise typer.BadParameter("Start date must be before or equal to end date.")
    ts = datetime.now()
    if output.exists():
        log.info(f"Output file {output} already exists. It will be overwritten.")
        output.unlink()  # Clean existing file

    with sync_playwright() as p:
        browser = p.webkit.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        current = start_date
        write_header = True
        while current <= end_date:
            next_month = (current.month % 12) + 1
            next_year = current.year + (current.month // 12)
            last_day = (datetime(next_year, next_month, 1) - timedelta(days=1)).day
            month_end = current.replace(day=last_day)
            if month_end > end_date:
                month_end = end_date

            fetch_and_export(
                page,
                current.strftime("%Y-%m-%d"),
                month_end.strftime("%Y-%m-%d"),
                csv_path=output,
                write_header=write_header,
                ts=ts,
            )
            write_header = False
            current = month_end + timedelta(days=1)

        browser.close()


if __name__ == "__main__":
    app()
